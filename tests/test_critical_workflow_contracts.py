from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook

from app_modules.export_identity import export_signature_for_state
from app_modules.pdf_artifact_state import current_pdf_artifact, store_pdf_artifact
from app_modules.project_identity import active_project_id_from_state
from app_modules.saved_project_builder import build_saved_project_from_state
from app_modules.saved_project_current_state import refresh_active_saved_project_current_snapshot
from app_modules.saved_project_serialization import saved_project_to_dict
from app_modules.workflow_state import ensure_workflow_defaults, mark_pdf_dirty
from calculator.calculator_state import CalculatorState
from calculator.formula_map import expected_row_formulas
from calculator.row_model import CalculatorRow
from calculator.workbook_export import export_calculation_workbook
from app_modules.parse_workflow import parse_and_normalize_itinerary


def _minimal_generated_state() -> dict[str, object]:
    state: dict[str, object] = {}
    ensure_workflow_defaults(state)
    state.update(
        {
            "itinerary_name": "Critical Oslo",
            "last_generated_raw_text": "Day 1\tHotel\t01/10/2026\t02/10/2026\tOslo\tHotel Bristol",
            "parsed_rows": [
                {
                    "day": "Day 1",
                    "type": "Hotel",
                    "city": "Oslo",
                    "title": "Hotel Bristol",
                    "start_date": "01/10/2026",
                    "end_date": "02/10/2026",
                }
            ],
            "output_edits": {
                "detail_level": "Rich descriptive",
                "output_brand": "agent",
                "days": {"Day 1": {"title": "Arrival in Oslo", "city": "Oslo"}},
            },
            "preview_signature": "preview-1",
            "itinerary_html": "<html><body>Arrival in Oslo</body></html>",
        }
    )
    return state


def test_messy_input_can_reach_parser_contract_without_ui() -> None:
    messy = """
    Day 1\tTransfer\t01/10/2026\t\t\t\tOslo\tPrivate airport to hotel transfer
    Day 1\tHotel\t01/10/2026\t02/10/2026\t\t\tOslo\t4 Star, Hotel Bristol, 1xNight, Incl Breakfast
    """.strip()

    rows = parse_and_normalize_itinerary(messy)

    assert len(rows) >= 2
    assert any(str(row.get("type", "")).lower() == "hotel" for row in rows)
    assert any("Oslo" in str(row.get("city") or row.get("description") or row) for row in rows)


def test_calculator_to_excel_contract_creates_real_workbook_bytes() -> None:
    state = CalculatorState(
        itinerary_name="Critical Calculator",
        rows=(
            CalculatorRow(
                row_id="1",
                day="Day 1",
                type="Hotel",
                travel_element="Hotel Bristol",
                gross_price_per_unit=100,
                units=2,
                supplier_currency="EUR",
                sales_currency="NOK",
            ),
        ),
    )

    export = export_calculation_workbook(state)
    workbook = load_workbook(BytesIO(export.content), data_only=False)

    assert export.filename == "Critical Calculator - Calculation.xlsx"
    assert workbook.sheetnames == ["Curr", "Kalk"]
    assert workbook["Kalk"]["J7"].value == "Hotel Bristol"
    assert workbook["Kalk"]["S7"].value == expected_row_formulas(7)["S"]


def test_saved_project_round_trip_preserves_identity_images_export_and_calculator_state() -> None:
    state = _minimal_generated_state()
    state["output_edits"] = {
        **state["output_edits"],
        "pictures_added": True,
        "day_images": {"Day 1": {"path": "images/oslo.jpg", "crop_focus": "center"}},
    }
    state["calculator_state"] = CalculatorState(
        itinerary_name="Critical Oslo",
        rows=(CalculatorRow(row_id="1", travel_element="Hotel Bristol"),),
    )

    project = build_saved_project_from_state(state, itinerary_name="Critical Oslo", project_id="project-1")
    payload = saved_project_to_dict(project)
    state["active_saved_project"] = payload

    assert refresh_active_saved_project_current_snapshot(state) is True
    assert active_project_id_from_state(state) == "project-1"
    assert state["active_saved_project"]["image_state"]["day_images"]["Day 1"]["path"] == "images/oslo.jpg"
    assert state["active_saved_project"]["calculator_snapshot"]["rows"][0]["travel_element"] == "Hotel Bristol"


def test_pdf_identity_changes_when_export_critical_image_state_changes() -> None:
    state = _minimal_generated_state()
    state["output_edits"] = {
        **state["output_edits"],
        "pictures_added": True,
        "day_images": {"Day 1": {"path": "images/oslo-a.jpg", "crop_focus": "top"}},
    }
    signature_a = export_signature_for_state(state)
    store_pdf_artifact(state, content=b"old-pdf", signature=signature_a, filename="old.pdf")

    state["output_edits"]["day_images"]["Day 1"] = {"path": "images/oslo-b.jpg", "crop_focus": "center"}

    assert export_signature_for_state(state) != signature_a
    assert current_pdf_artifact(state) is None

    mark_pdf_dirty(state)
    assert state["pdf_status"] == "Needs refresh"
    assert state["pdf_bytes"] is None


def test_pdf_crop_focus_read_is_pure_and_does_not_create_day_image_defaults() -> None:
    from app_modules.export_render_context import day_image_crop_focus_for_grouped_days

    output_edits = {"day_images": {"Day 1": {"path": "images/oslo.jpg", "crop_focus": "center"}}}

    focus = day_image_crop_focus_for_grouped_days({"Day 1": [], "Day 2": []}, output_edits)

    assert focus == {"Day 1": "center", "Day 2": "top"}
    assert output_edits == {"day_images": {"Day 1": {"path": "images/oslo.jpg", "crop_focus": "center"}}}


def test_pdf_identity_changes_when_export_layout_or_cover_state_changes() -> None:
    state = _minimal_generated_state()
    state["output_edits"] = {
        **state["output_edits"],
        "cover_image": {"path": "images/cover-a.jpg", "crop_focus": "top"},
        "day_page_layout": "narrative",
    }
    signature_a = export_signature_for_state(state)

    state["output_edits"]["cover_image"] = {"path": "images/cover-a.jpg", "crop_focus": "bottom"}
    signature_b = export_signature_for_state(state)
    state["output_edits"]["day_page_layout"] = "compact"

    assert signature_a != signature_b
    assert signature_b != export_signature_for_state(state)
