from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook as openpyxl_load_workbook

import calculator.library_workbook as library_workbook_module
import calculator.library_store as library_store_module
from calculator.library_store import LocalLibraryStore
from calculator.library_workbook import (
    DATA_SHEETS,
    LocalLibraryDiagnostic,
    LocalLibraryWorkbookError,
    clear_local_library_workbook_cache,
    load_local_library_workbook,
)

_HEADERS = (
    "ID",
    "Type",
    "Travel element",
    "Gross P per unit",
    "Supp Comm",
    "Supp curr",
    "Sales P per unit",
    "Sales curr",
)


def _write_workbook(
    path: Path,
    *,
    rows_by_sheet: dict[str, list[tuple[object, ...]]] | None = None,
    currency_rows: list[tuple[object, object]] | None = None,
    omit_sheet: str = "",
    headers_by_sheet: dict[str, tuple[str, ...]] | None = None,
) -> Path:
    workbook = Workbook()
    workbook.remove(workbook.active)
    rows_by_sheet = rows_by_sheet or {}
    headers_by_sheet = headers_by_sheet or {}

    if omit_sheet != "Curr":
        currency_sheet = workbook.create_sheet("Curr")
        currency_sheet.append((None, None, "NOK"))
        for code, rate in currency_rows or [("NOK", 1), ("EUR", 11)]:
            currency_sheet.append((None, code, rate))

    for sheet_name in DATA_SHEETS:
        if sheet_name == omit_sheet:
            continue
        sheet = workbook.create_sheet(sheet_name)
        sheet.append(headers_by_sheet.get(sheet_name, _HEADERS))
        for row in rows_by_sheet.get(sheet_name, []):
            sheet.append(row)

    workbook.save(path)
    workbook.close()
    return path


def _valid_row(*, travel_element: str = "Oslo: Valid hotel", currency: str = "NOK") -> tuple[object, ...]:
    return ("NO", "Hotel", travel_element, 100, 0, currency, 120, currency)


def test_malformed_numeric_row_is_skipped_without_becoming_zero(tmp_path: Path) -> None:
    path = _write_workbook(
        tmp_path / "malformed-number.xlsx",
        rows_by_sheet={
            "General": [_valid_row()],
            "Hotels": [("NO", "Hotel", "Oslo: Broken price", "not-a-number", 0, "NOK", 120, "NOK")],
        },
    )

    library = load_local_library_workbook(path)

    assert [row.travel_element for row in library.rows] == ["Oslo: Valid hotel"]
    assert all(row.travel_element != "Oslo: Broken price" for row in library.rows)
    issue = library.invalid_records[0]
    assert issue.code == "invalid_numeric_value"
    assert issue.worksheet == "Hotels"
    assert issue.excel_row == 2
    assert issue.field == "Gross P per unit"
    assert issue.value == "'not-a-number'"
    assert "unsupported characters" in issue.message


def test_unsupported_currency_only_skips_affected_row(tmp_path: Path) -> None:
    path = _write_workbook(
        tmp_path / "unsupported-currency.xlsx",
        rows_by_sheet={
            "General": [_valid_row()],
            "Hotels": [_valid_row(travel_element="Oslo: Unsupported", currency="XYZ")],
        },
    )

    library = load_local_library_workbook(path)

    assert [row.travel_element for row in library.rows] == ["Oslo: Valid hotel"]
    issue = library.invalid_records[0]
    assert issue.code == "unsupported_currency"
    assert issue.worksheet == "Hotels"
    assert issue.field == "Supp curr"
    assert issue.value == "'XYZ'"


def test_invalid_currency_rate_is_warning_and_valid_nok_rows_still_load(tmp_path: Path) -> None:
    path = _write_workbook(
        tmp_path / "invalid-rate.xlsx",
        currency_rows=[("NOK", 1), ("EUR", "bad-rate")],
        rows_by_sheet={"General": [_valid_row()]},
    )

    library = load_local_library_workbook(path)

    assert len(library.rows) == 1
    assert library.currency_rates == {"NOK": 1.0}
    assert library.invalid_records == ()
    warning = library.warnings[0]
    assert warning.code == "invalid_currency_rate"
    assert warning.worksheet == "Curr"
    assert warning.excel_row == 3
    assert warning.field == "Rate"
    assert warning.value == "'bad-rate'"


def test_invalid_nok_rate_does_not_remove_fixed_base_currency(tmp_path: Path) -> None:
    path = _write_workbook(
        tmp_path / "invalid-nok-rate.xlsx",
        currency_rows=[("NOK", "bad-rate"), ("EUR", 11)],
        rows_by_sheet={"General": [_valid_row()]},
    )

    library = load_local_library_workbook(path)

    assert len(library.rows) == 1
    assert library.currency_rates["NOK"] == 1.0
    assert library.warnings[0].code == "invalid_currency_rate"


def test_formula_without_cached_value_skips_only_that_row(tmp_path: Path) -> None:
    path = _write_workbook(
        tmp_path / "missing-formula-cache.xlsx",
        rows_by_sheet={
            "General": [_valid_row()],
            "Hotels": [("NO", "Hotel", "Oslo: Formula price", "=50+50", 0, "NOK", 120, "NOK")],
        },
    )

    library = load_local_library_workbook(path)

    assert [row.travel_element for row in library.rows] == ["Oslo: Valid hotel"]
    issue = library.invalid_records[0]
    assert issue.code == "invalid_formula_cache"
    assert issue.worksheet == "Hotels"
    assert issue.excel_row == 2
    assert issue.field == "Gross P per unit"
    assert "no cached value" in issue.message


@pytest.mark.parametrize(
    ("omit_sheet", "expected_code"),
    [("Activities", "missing_sheets")],
)
def test_missing_required_sheet_is_schema_error(tmp_path: Path, omit_sheet: str, expected_code: str) -> None:
    path = _write_workbook(tmp_path / "missing-sheet.xlsx", omit_sheet=omit_sheet)

    with pytest.raises(LocalLibraryWorkbookError) as caught:
        load_local_library_workbook(path)

    assert caught.value.category == "schema"
    assert caught.value.code == expected_code
    assert omit_sheet in str(caught.value)


def test_missing_required_header_is_schema_error(tmp_path: Path) -> None:
    headers = tuple(header for header in _HEADERS if header != "Sales curr")
    path = _write_workbook(
        tmp_path / "missing-header.xlsx",
        headers_by_sheet={"Hotels": headers},
        rows_by_sheet={"General": [_valid_row()]},
    )

    with pytest.raises(LocalLibraryWorkbookError) as caught:
        load_local_library_workbook(path)

    assert caught.value.category == "schema"
    assert caught.value.code == "missing_headers"
    assert "Hotels" in str(caught.value)
    assert "Sales curr" in str(caught.value)


def test_all_invalid_rows_keep_actionable_diagnostics_on_fatal_result(tmp_path: Path) -> None:
    path = _write_workbook(
        tmp_path / "all-invalid.xlsx",
        rows_by_sheet={
            "Hotels": [("NO", "Hotel", "Oslo: Broken", "bad-price", 0, "NOK", 120, "NOK")],
        },
    )

    with pytest.raises(LocalLibraryWorkbookError) as caught:
        load_local_library_workbook(path)

    assert caught.value.category == "fatal_workbook"
    assert caught.value.code == "no_fetchable_rows"
    assert len(caught.value.diagnostics) == 1
    assert caught.value.diagnostics[0].worksheet == "Hotels"
    assert caught.value.diagnostics[0].field == "Gross P per unit"


def test_unreadable_xlsx_is_fatal_workbook_error(tmp_path: Path) -> None:
    path = tmp_path / "unreadable.xlsx"
    path.write_bytes(b"not an xlsx package")

    with pytest.raises(LocalLibraryWorkbookError) as caught:
        load_local_library_workbook(path)

    assert caught.value.category == "fatal_workbook"
    assert caught.value.code == "unreadable_workbook"


def test_store_preserves_error_category_and_row_diagnostics(monkeypatch: pytest.MonkeyPatch) -> None:
    diagnostic = LocalLibraryDiagnostic(
        "invalid_record",
        "invalid_numeric_value",
        "Hotels row 9, Gross P per unit: 'bad'. Invalid number.",
        "Hotels",
        9,
        "Gross P per unit",
        "'bad'",
    )

    def fail_load() -> object:
        raise LocalLibraryWorkbookError(
            "Local Library workbook contains no fetchable rows.",
            category="fatal_workbook",
            code="no_fetchable_rows",
            diagnostics=(diagnostic,),
        )

    monkeypatch.setattr(library_store_module, "load_local_library_authority", fail_load)

    result = LocalLibraryStore().list_rows()

    assert result.error_category == "fatal_workbook"
    assert result.diagnostics == (diagnostic,)


def test_loader_closes_cached_value_workbook(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    path = _write_workbook(
        tmp_path / "close-resources.xlsx",
        rows_by_sheet={"General": [_valid_row()]},
    )
    closed: list[bool] = []

    class TrackingWorkbook:
        def __init__(self, workbook: object) -> None:
            self._workbook = workbook

        def __getattr__(self, name: str) -> object:
            return getattr(self._workbook, name)

        def __getitem__(self, key: str) -> object:
            return self._workbook[key]  # type: ignore[index]

        def close(self) -> None:
            closed.append(True)
            self._workbook.close()  # type: ignore[attr-defined]

    def tracking_load_workbook(*args: object, **kwargs: object) -> TrackingWorkbook:
        return TrackingWorkbook(openpyxl_load_workbook(*args, **kwargs))

    clear_local_library_workbook_cache()
    monkeypatch.setattr(library_workbook_module, "load_workbook", tracking_load_workbook)

    library = load_local_library_workbook(path)

    assert len(library.rows) == 1
    assert closed == [True]
