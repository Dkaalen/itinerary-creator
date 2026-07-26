from __future__ import annotations

import os
from pathlib import Path
from types import MappingProxyType

import pytest
from openpyxl import Workbook

from calculator.library_authority import (
    LOCAL_LIBRARY_AUTHORITY_ID,
    clear_local_library_authority_cache,
    load_local_library_authority,
    local_library_authority_path,
    local_library_supported_worksheets,
)
from calculator.library_normalize import normalize_library_mapping
from calculator.library_workbook_models import LocalLibraryWorkbookError

_HEADERS = (
    "ID", "Type", "Travel element", "Gross P per unit", "Supp Comm",
    "Supp curr", "Sales P per unit", "Sales curr",
)


def _workbook(path: Path, title: str = "Oslo hotel") -> Path:
    workbook = Workbook()
    workbook.remove(workbook.active)
    curr = workbook.create_sheet("Curr")
    curr.append((None, None, "NOK"))
    curr.append((None, "NOK", 1))
    curr.append((None, "EUR", 11))
    for sheet_name in ("General", "Hotels", "Transfers", "Transport", "Activities"):
        sheet = workbook.create_sheet(sheet_name)
        sheet.append(_HEADERS)
        if sheet_name == "General":
            sheet.append(("NO", "Hotel", title, 100, 0, "NOK", 120, "NOK"))
    workbook.save(path)
    workbook.close()
    return path


def test_authority_exposes_path_fingerprint_worksheets_records_and_identity() -> None:
    clear_local_library_authority_cache()
    authority = load_local_library_authority()

    assert authority.authority_id == LOCAL_LIBRARY_AUTHORITY_ID
    assert authority.path == local_library_authority_path().resolve()
    assert authority.fingerprint.startswith("xlsx-sha256-v1:")
    assert authority.supported_worksheets == local_library_supported_worksheets()
    assert authority.supported_worksheets == ("Curr", "General", "Hotels", "Transfers", "Transport", "Activities")
    assert len(authority.records) == 5966
    assert len({row.source_identity for row in authority.records}) == len(authority.records)
    assert len({row.library_id for row in authority.records}) == len(authority.records)


def test_source_identity_uses_workbook_sheet_and_row_not_display_text() -> None:
    base = {
        "source_workbook": "library.xlsx", "source_sheet": "Activities", "source_row": 19,
        "ID": "FI", "Type": "Activity", "Travel element": "First title",
    }
    renamed = {**base, "Travel element": "Completely different title", "Supplier": "Changed supplier"}
    other_row = {**renamed, "source_row": 20}
    other_sheet = {**renamed, "source_sheet": "Transfers"}

    first = normalize_library_mapping(base)
    assert first.library_id == normalize_library_mapping(renamed).library_id
    assert first.source_identity == normalize_library_mapping(renamed).source_identity
    assert first.library_id != normalize_library_mapping(other_row).library_id
    assert first.library_id != normalize_library_mapping(other_sheet).library_id


def test_intentional_identical_workbook_products_remain_separate() -> None:
    authority = load_local_library_authority()
    rows = [row for row in authority.records if row.source_sheet == "Activities" and row.source_row in {19, 20}]
    assert len(rows) == 2
    assert rows[0].travel_element == rows[1].travel_element
    assert rows[0].library_id != rows[1].library_id
    assert rows[0].source_identity != rows[1].source_identity


def test_cache_reuses_unchanged_content_and_reports_explicit_clear(tmp_path: Path) -> None:
    path = _workbook(tmp_path / "library.xlsx")
    clear_local_library_authority_cache(path)
    first = load_local_library_authority(path)
    second = load_local_library_authority(path)

    assert first.cache_status == "miss"
    assert first.cache_invalidation_reason == "explicit_cache_clear"
    assert second.cache_status == "hit"
    assert second.cache_invalidation_reason == "unchanged"
    assert first.records is second.records


def test_content_change_invalidates_even_when_mtime_is_restored(tmp_path: Path) -> None:
    path = _workbook(tmp_path / "library.xlsx", "Oslo hotel")
    clear_local_library_authority_cache(path)
    first = load_local_library_authority(path)
    original_stat = path.stat()

    _workbook(path, "Bergen hotel")
    os.utime(path, ns=(original_stat.st_atime_ns, original_stat.st_mtime_ns))
    changed = load_local_library_authority(path)

    assert changed.fingerprint != first.fingerprint
    assert changed.cache_status == "miss"
    assert changed.cache_invalidation_reason == "workbook_content_changed"
    assert changed.records[0].travel_element == "Bergen hotel"


def test_failed_read_does_not_poison_next_successful_read(tmp_path: Path) -> None:
    path = tmp_path / "library.xlsx"
    path.write_bytes(b"not an xlsx")
    clear_local_library_authority_cache(path)
    with pytest.raises(LocalLibraryWorkbookError):
        load_local_library_authority(path)

    _workbook(path)
    recovered = load_local_library_authority(path)
    assert recovered.records[0].travel_element == "Oslo hotel"


def test_cached_records_and_rates_cannot_be_mutated() -> None:
    authority = load_local_library_authority()
    assert isinstance(authority.currency_rates, MappingProxyType)
    with pytest.raises(TypeError):
        authority.currency_rates["EUR"] = 99  # type: ignore[index]
    with pytest.raises(Exception):
        authority.records[0].travel_element = "mutated"  # type: ignore[misc]


def test_production_authority_has_no_alternate_storage_backend() -> None:
    production_files = (
        Path("calculator/library_authority.py"),
        Path("calculator/library_store.py"),
        Path("calculator/library_workbook.py"),
    )
    text = "\n".join(path.read_text(encoding="utf-8").lower() for path in production_files)
    forbidden = ("supabase", "sqlite", "gspread", "google sheets", "requests.get", "httpx", "read_csv", "json.load")
    assert not [token for token in forbidden if token in text]


def test_failed_changed_read_does_not_replace_last_known_good_snapshot(tmp_path: Path) -> None:
    path = _workbook(tmp_path / "library.xlsx", "Known good")
    clear_local_library_authority_cache(path)
    good = load_local_library_authority(path)

    path.write_bytes(b"corrupt replacement")
    with pytest.raises(LocalLibraryWorkbookError):
        load_local_library_authority(path)

    _workbook(path, "Recovered workbook")
    recovered = load_local_library_authority(path)
    assert good.records[0].travel_element == "Known good"
    assert recovered.records[0].travel_element == "Recovered workbook"
    assert recovered.fingerprint != good.fingerprint
