from pathlib import Path

from openpyxl import load_workbook
from calculator.library_store import LocalLibraryStore
from calculator.library_workbook import (
    DATA_SHEETS,
    REQUIRED_SHEETS,
    WORKBOOK_PATH,
    _has_product_content,
    _headers,
    load_local_library_workbook,
)

def test_bundled_workbook_is_authoritative_and_complete():
    library = load_local_library_workbook()
    assert WORKBOOK_PATH.is_file()
    workbook = load_workbook(WORKBOOK_PATH, read_only=True, data_only=True)
    try:
        source_product_rows = 0
        for sheet_name in DATA_SHEETS:
            sheet = workbook[sheet_name]
            headers, header_row = _headers(sheet)
            for cells in sheet.iter_rows(min_row=header_row + 1, values_only=True):
                mapping = {header: value for header, value in zip(headers, cells) if header}
                source_product_rows += int(_has_product_content(mapping))
    finally:
        workbook.close()

    invalid_source_rows = {
        (diagnostic.worksheet, diagnostic.excel_row)
        for diagnostic in library.invalid_records
        if diagnostic.worksheet and diagnostic.excel_row is not None
    }
    assert len(library.rows) + len(invalid_source_rows) == source_product_rows
    assert {row.source_sheet for row in library.rows} == set(REQUIRED_SHEETS) - {"Curr"}
    assert library.currency_rates["EUR"] == 11.0
    assert library.currency_rates["NOK"] == 1.0

def test_store_uses_only_local_excel():
    result = LocalLibraryStore().list_rows()
    assert result.source == "local_excel"
    assert result.read_only is True
    assert result.message == ""
    assert result.rows

def test_runtime_store_has_no_google_or_supabase_route():
    source = Path("calculator/library_store.py").read_text(encoding="utf-8").lower()
    assert "google" not in source
    assert "gspread" not in source
    assert "supabase" not in source
    assert "fallback" not in source


def test_intentional_duplicate_rows_are_preserved_with_distinct_source_identity():
    library = load_local_library_workbook()
    grouped: dict[tuple[object, ...], list[object]] = {}
    for row in library.rows:
        key = (row.source_sheet, row.travel_element, row.supplier, row.gross_price_per_unit, row.supplier_currency)
        grouped.setdefault(key, []).append(row)
    duplicate_group = next(group for group in grouped.values() if len(group) > 1)
    assert len({row.library_id for row in duplicate_group}) == len(duplicate_group)
    assert len({row.source_row for row in duplicate_group}) == len(duplicate_group)
