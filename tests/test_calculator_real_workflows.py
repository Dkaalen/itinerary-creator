from __future__ import annotations

from io import BytesIO

import pytest
from openpyxl import load_workbook

from calculator.calculations import calculate_dashboard
from calculator.calculator_state import CalculatorState
from calculator.columns import DATA_START_ROW
from calculator.row_model import CalculatorRow
from calculator.state_serialization import calculator_state_from_json, calculator_state_to_json
from calculator.workbook_export import export_calculation_workbook


def _city_break() -> tuple[CalculatorState, dict[str, float]]:
    return (
        CalculatorState(
            itinerary_name="Oslo City Break",
            number_of_pax=8,
            rows=(
                CalculatorRow(row_id="1", day="Day 1", type="Transfer", travel_element="Airport transfer", gross_price_per_unit=2400, units=1, supplier_currency="NOK", sales_price_per_unit=3000, sales_currency="NOK", vat12=321.43),
                CalculatorRow(row_id="2", day="Day 1", type="Hotel", travel_element="Central hotel", gross_price_per_unit=155.75, units=8, supplier_commission=0.10, supplier_currency="EUR", sales_price_per_unit=215, sales_currency="EUR", vat12=180),
                CalculatorRow(row_id="3", day="Day 2", type="Activity", travel_element="Fjord cruise", gross_price_per_unit="=42.5*0.9", units=8, supplier_currency="EUR", sales_price_per_unit=62.5, sales_currency="EUR", vat0_international=200),
            ),
        ),
        {"NOK": 1, "EUR": 11.72},
    )


def _self_drive() -> tuple[CalculatorState, dict[str, float]]:
    rows = []
    currencies = ("NOK", "EUR", "USD")
    for index in range(40):
        currency = currencies[index % len(currencies)]
        rows.append(
            CalculatorRow(
                row_id=str(index + 1),
                day=f"Day {index // 3 + 1}",
                type="Hotel" if index % 3 == 0 else "Activity",
                travel_element=f"Service {index + 1}",
                gross_price_per_unit=75.005 + index,
                units=(index % 4) + 1,
                supplier_commission=0.125 if index % 5 == 0 else 0,
                supplier_currency=currency,
                sales_price_per_unit=110.25 + index,
                sales_currency="NOK" if index % 2 == 0 else currency,
            )
        )
    return CalculatorState(itinerary_name="Long Self Drive", number_of_pax=17, rows=tuple(rows)), {"NOK": 1, "EUR": 11.85, "USD": 10.15}


def _manual_overrides() -> tuple[CalculatorState, dict[str, float]]:
    return (
        CalculatorState(
            itinerary_name="Manual Overrides",
            rows=(
                CalculatorRow(
                    row_id="1",
                    travel_element="Contracted hotel",
                    gross_price_per_unit=100,
                    units=12,
                    supplier_commission=0.15,
                    supplier_currency="EUR",
                    sales_currency="EUR",
                    gross_price_override=1199.99,
                    net_price_override=999.95,
                    supplier_x_rate_override=11.9,
                    sales_price_nok_total_override=18000,
                    gp_nok_override=6100.6,
                ),
                CalculatorRow(row_id="2", travel_element="Price pending", gross_price_per_unit="", units=1, supplier_currency="NOK", sales_currency="NOK"),
            ),
        ),
        {"NOK": 1, "EUR": 11.9},
    )


@pytest.mark.parametrize("scenario", [_city_break, _self_drive, _manual_overrides], ids=["city-break", "self-drive", "manual-overrides"])
def test_realistic_calculator_workflows_survive_save_reload_and_excel_export(scenario) -> None:
    state, rates = scenario()
    expected_dashboard = calculate_dashboard(state.rows, state.number_of_pax, rates)

    restored = calculator_state_from_json(calculator_state_to_json(state))
    assert restored == state
    assert calculate_dashboard(restored.rows, restored.number_of_pax, rates) == expected_dashboard

    export = export_calculation_workbook(restored, currency_rates=rates)
    workbook = load_workbook(BytesIO(export.content), data_only=False)
    sheet = workbook["Kalk"]

    assert sheet[f"J{DATA_START_ROW}"].value == state.rows[0].travel_element
    assert sheet[f"J{DATA_START_ROW + len(state.rows) - 1}"].value == state.rows[-1].travel_element
    assert sheet[f"Z{DATA_START_ROW}"].value is not None
    assert workbook.calculation.fullCalcOnLoad is True
    assert workbook.calculation.forceFullCalc is True


def test_dashboard_passenger_count_is_display_only() -> None:
    state, rates = _city_break()
    without_pax = calculate_dashboard(state.rows, None, rates)
    with_pax = calculate_dashboard(state.rows, 8, rates)

    assert with_pax.total_cost_nok == without_pax.total_cost_nok
    assert with_pax.total_sales_nok == without_pax.total_sales_nok
    assert with_pax.cost_per_pax == pytest.approx(with_pax.total_cost_nok / 8, abs=0.01)
    assert with_pax.sales_per_pax == pytest.approx(with_pax.total_sales_nok / 8, abs=0.01)


def test_project_switch_replaces_calculator_rows_rates_and_dashboard_without_leakage() -> None:
    from app_modules.calculator_state_keys import CALCULATOR_STATE_KEY, CURRENCY_RATES_STATE_KEY
    from app_modules.saved_project_calculator_state import (
        apply_calculator_snapshot_to_state,
        calculator_snapshot_from_workflow_state,
    )

    city_state, city_rates = _city_break()
    self_drive_state, self_drive_rates = _self_drive()
    city_snapshot = calculator_snapshot_from_workflow_state(
        {
            CALCULATOR_STATE_KEY: city_state,
            CURRENCY_RATES_STATE_KEY: city_rates,
        }
    )
    session = {
        CALCULATOR_STATE_KEY: self_drive_state,
        CURRENCY_RATES_STATE_KEY: self_drive_rates,
    }

    apply_calculator_snapshot_to_state(session, city_snapshot)

    restored = session[CALCULATOR_STATE_KEY]
    assert restored == city_state
    assert session[CURRENCY_RATES_STATE_KEY] == city_snapshot["currency_rates"]
    assert session[CURRENCY_RATES_STATE_KEY]["EUR"] == city_rates["EUR"]
    assert session[CURRENCY_RATES_STATE_KEY]["USD"] != self_drive_rates["USD"]
    assert all(row.travel_element != "Service 40" for row in restored.rows)
    assert calculate_dashboard(restored.rows, restored.number_of_pax, session[CURRENCY_RATES_STATE_KEY]) == calculate_dashboard(
        city_state.rows,
        city_state.number_of_pax,
        city_snapshot["currency_rates"],
    )

    export = export_calculation_workbook(restored, currency_rates=session[CURRENCY_RATES_STATE_KEY])
    workbook = load_workbook(BytesIO(export.content), data_only=False)
    assert workbook["Kalk"][f"J{DATA_START_ROW}"].value == "Airport transfer"
    assert workbook["Kalk"][f"J{DATA_START_ROW + 2}"].value == "Fjord cruise"
