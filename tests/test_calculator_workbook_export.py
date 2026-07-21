from __future__ import annotations

from io import BytesIO
from pathlib import Path
import re
from time import perf_counter
from zipfile import ZipFile

import pytest
from openpyxl import load_workbook

from calculator.calculator_state import CalculatorState
from calculator.columns import ADVANCED_COLUMN_RANGES, DATA_END_ROW, DATA_START_ROW
from calculator.formula_map import PAYMENT_FORMULAS, TOTAL_FORMULAS, expected_row_formulas
from calculator.filename_sanitizer import calculation_workbook_filename, sanitize_filename_stem
from calculator.row_model import CalculatorRow
from calculator.template_structure import inspect_template_structure
from calculator.workbook_export import (
    build_calculation_workbook,
    export_calculation_workbook,
    save_calculation_workbook,
)


def test_filename_sanitizer_creates_windows_safe_calculation_name() -> None:
    assert sanitize_filename_stem(' Tromsø: Northern/ Lights? 2026. ') == "Tromsø Northern Lights 2026"
    assert sanitize_filename_stem("   ") == "Itinerary"
    assert calculation_workbook_filename("Tromsø Northern Lights 2026") == "Tromsø Northern Lights 2026 - Calculation.xlsx"


def test_build_workbook_fills_rows_and_preserves_core_formulas() -> None:
    state = CalculatorState(
        itinerary_name="Tromsø Northern Lights 2026",
        rows=(
            CalculatorRow(
                row_id="1",
                day="Day 1",
                type="Hotel",
                from_date="01/10/2026",
                to_date="02/10/2026",
                supplier="Scandic",
                travel_element="Scandic Simonkenttä",
                manual_booking=True,
                status="Booked",
                comments="Breakfast included",
                non_refundable=True,
                refundable=False,
                url="https://example.com/hotel",
                gross_price_per_unit=100,
                units=2,
                supplier_commission=0.1,
                supplier_currency="eur",
                sales_price_per_unit=150,
                sales_currency="nok",
                vat25=10,
                vat15=1,
            ),
            CalculatorRow(
                row_id="2",
                day="Day 2",
                type="Activity",
                travel_element="Northern lights tour",
                gross_price_per_unit=50,
                units=4,
                supplier_commission=0,
                supplier_currency="NOK",
                sales_currency="NOK",
                vat12=2,
                vat0_domestic=3,
                vat0_international=4,
            ),
        ),
    )

    workbook = build_calculation_workbook(state)
    sheet = workbook["Kalk"]

    assert sheet["B7"].value == "1"
    assert sheet["C7"].value == "Day 1"
    assert sheet["D7"].value == "Hotel"
    assert sheet["E7"].value == "01/10/2026"
    assert sheet["F7"].value == "02/10/2026"
    assert sheet["I7"].value == "Scandic"
    assert sheet["J7"].value == "Scandic Simonkenttä"
    assert sheet["K7"].value is True
    assert sheet["L7"].value == "Booked"
    assert sheet["M7"].value == "Breakfast included"
    assert sheet["N7"].value is True
    assert sheet["O7"].value is False
    assert sheet["P7"].value == "https://example.com/hotel"
    assert sheet["Q7"].value == 100
    assert sheet["R7"].value == 2
    assert sheet["T7"].value == 0.1
    assert sheet["V7"].value == "EUR"
    assert sheet["Y7"].value == 150
    assert sheet["AA7"].value == "NOK"
    assert sheet["AF7"].value == 10
    assert sheet["AG7"].value == 1

    assert sheet["B8"].value == "2"
    assert sheet["J8"].value == "Northern lights tour"
    assert sheet["Y8"].value == "=IFERROR(Q8*W8/AB8,0)"

    expected = expected_row_formulas(7)
    for column, formula in expected.items():
        if column == "Y":
            continue
        assert sheet[f"{column}7"].value == formula

    assert sheet["Z101"].value == TOTAL_FORMULAS["Z101"]
    assert sheet["AC101"].value == TOTAL_FORMULAS["AC101"]
    assert sheet["Z103"].value == "=Z101"
    assert sheet["Y104"].value is None
    assert sheet["Z104"].value is None


def test_exported_workbook_preserves_template_structure_and_styles() -> None:
    state = CalculatorState(
        itinerary_name="Export Test",
        rows=(CalculatorRow(row_id="1", travel_element="Styled line", gross_price_per_unit=10, units=1),),
    )

    export = export_calculation_workbook(state)
    workbook = load_workbook(BytesIO(export.content), data_only=False)
    sheet = workbook["Kalk"]

    assert export.filename == "Export Test - Calculation.xlsx"
    assert workbook.sheetnames == ["Curr", "Kalk"]
    assert sheet.auto_filter.ref == "B6:AE101"
    assert inspect_template_structure().hidden_column_ranges == ADVANCED_COLUMN_RANGES
    assert sheet.column_dimensions["G"].hidden is True
    assert sheet.column_dimensions["G"].outlineLevel == 1
    assert sheet.column_dimensions["J"].collapsed is True
    assert sheet.column_dimensions["P"].collapsed is True
    assert sheet.sheet_view.showOutlineSymbols is not False
    assert sheet["B6"].value == "ID"
    assert sheet["B6"].fill.fill_type is not None
    assert sheet["S7"].value == "=ROUND(Q7*R7,2)"


def test_save_calculation_workbook_writes_xlsx_file(tmp_path) -> None:
    state = CalculatorState(
        itinerary_name="Oslo Fjord",
        rows=(CalculatorRow(row_id="1", travel_element="Fjord cruise", gross_price_per_unit=10, units=2),),
    )

    output_path = save_calculation_workbook(state, tmp_path)

    assert output_path.name == "Oslo Fjord - Calculation.xlsx"
    workbook = load_workbook(output_path, data_only=False)
    assert workbook["Kalk"]["J7"].value == "Fjord cruise"


def test_export_rejects_more_rows_than_template_can_hold() -> None:
    rows = tuple(CalculatorRow(row_id=str(index)) for index in range(DATA_START_ROW, DATA_END_ROW + 2))

    with pytest.raises(ValueError, match="at most 93 rows"):
        build_calculation_workbook(CalculatorState(rows=rows))


def test_exported_workbook_writes_default_currency_table_and_xrates() -> None:
    state = CalculatorState(
        itinerary_name="Currency Test",
        rows=(
            CalculatorRow(row_id="1", gross_price_per_unit=100, units=1, supplier_currency="EUR", sales_currency="USD"),
            CalculatorRow(row_id="2", gross_price_per_unit=100, units=1, supplier_currency="NOK", sales_currency="GBP"),
        ),
    )

    workbook = build_calculation_workbook(state)
    curr = workbook["Curr"]
    sheet = workbook["Kalk"]

    assert curr["B2"].value == "NOK"
    assert curr["C2"].value == 1
    assert curr["B3"].value == "EUR"
    assert curr["C3"].value == 11
    assert curr["B4"].value == "USD"
    assert curr["C4"].value == 10
    assert sheet["W7"].value == expected_row_formulas(7)["W"]
    assert sheet["AB7"].value == expected_row_formulas(7)["AB"]
    assert sheet["W8"].value == expected_row_formulas(8)["W"]
    assert sheet["AB8"].value == expected_row_formulas(8)["AB"]


def test_exported_workbook_uses_edited_currency_rates() -> None:
    state = CalculatorState(
        itinerary_name="Edited Currency Test",
        rows=(CalculatorRow(row_id="1", gross_price_per_unit=100, units=1, supplier_currency="EUR", sales_currency="USD"),),
    )

    workbook = build_calculation_workbook(state, currency_rates={"EUR": 12.25, "USD": 9.75, "NOK": 1})
    curr = workbook["Curr"]
    sheet = workbook["Kalk"]

    assert curr["C2"].value == 1
    assert curr["C3"].value == 12.25
    assert curr["C4"].value == 9.75
    assert sheet["W7"].value == expected_row_formulas(7)["W"]
    assert sheet["AB7"].value == expected_row_formulas(7)["AB"]


def test_workbook_vat_totals_cover_every_data_row() -> None:
    workbook = build_calculation_workbook(
        CalculatorState(rows=(CalculatorRow(row_id="1", vat25=123, supplier_currency="NOK", sales_currency="NOK"),))
    )
    sheet = workbook["Kalk"]

    assert sheet["AF101"].value == "=ROUND(SUM(AF7:AF99),2)"
    assert sheet["AG101"].value == "=ROUND(SUM(AG7:AG99),2)"
    assert sheet["AJ101"].value == "=ROUND(SUM(AJ7:AJ99),2)"


def test_workbook_keeps_live_currency_lookups_unless_rate_is_overridden() -> None:
    state = CalculatorState(
        rows=(
            CalculatorRow(row_id="1", supplier_currency="EUR", sales_currency="USD"),
            CalculatorRow(
                row_id="2",
                supplier_currency="EUR",
                supplier_x_rate_override=12.34,
                sales_currency="USD",
                sales_x_rate_override=9.87,
            ),
        )
    )

    sheet = build_calculation_workbook(state)["Kalk"]

    assert sheet["W7"].value == expected_row_formulas(7)["W"]
    assert sheet["AB7"].value == expected_row_formulas(7)["AB"]
    assert sheet["W8"].value == 12.34
    assert sheet["AB8"].value == 9.87


def test_zero_sales_price_uses_same_gross_price_fallback_as_app() -> None:
    state = CalculatorState(
        rows=(
            CalculatorRow(
                row_id="1",
                gross_price_per_unit=125,
                units=2,
                sales_price_per_unit=0,
                supplier_currency="NOK",
                sales_currency="NOK",
            ),
        )
    )

    sheet = build_calculation_workbook(state)["Kalk"]

    assert sheet["Y7"].value == "=IFERROR(Q7*W7/AB7,0)"
    assert sheet["Z7"].value == expected_row_formulas(7)["Z"]


def test_download_export_completes_within_interactive_budget() -> None:
    state = CalculatorState(
        itinerary_name="Performance",
        rows=tuple(
            CalculatorRow(
                row_id=str(index),
                day=f"Day {index}",
                type="Hotel",
                travel_element=f"Hotel {index}",
                gross_price_per_unit=100 + index,
                units=2,
                supplier_currency="NOK",
                sales_currency="EUR",
            )
            for index in range(1, 31)
        ),
    )

    started = perf_counter()
    export = export_calculation_workbook(state)
    elapsed = perf_counter() - started

    assert export.content.startswith(b"PK")
    assert elapsed < 1.25

def test_download_export_preserves_reference_package_outside_approved_cells() -> None:
    state = CalculatorState(
        itinerary_name="Reference parity",
        rows=(
            CalculatorRow(
                row_id="1",
                day="Day 1",
                type="Hotel",
                supplier="Nordic Hotel",
                travel_element="Two nights",
                gross_price_per_unit=125.5,
                units=2,
                supplier_currency="EUR",
                sales_price_per_unit=175,
                sales_currency="NOK",
                vat25=25,
            ),
        ),
    )

    export = export_calculation_workbook(state)
    template_path = Path(__file__).parents[1] / "calculator" / "templates" / "Calculation-template-Mal.xlsx"
    with ZipFile(template_path) as reference, ZipFile(BytesIO(export.content)) as generated:
        assert generated.namelist() == reference.namelist()
        changed_parts = {
            name for name in reference.namelist() if reference.read(name) != generated.read(name)
        }
        assert changed_parts == {
            "xl/worksheets/sheet1.xml",
            "xl/worksheets/sheet2.xml",
            "xl/workbook.xml",
        }
        for name in reference.namelist():
            if name not in changed_parts:
                assert generated.read(name) == reference.read(name)

        generated_workbook = generated.read("xl/workbook.xml").decode("utf-8")
        assert 'calcMode="auto"' in generated_workbook
        assert 'fullCalcOnLoad="1"' in generated_workbook
        assert 'forceFullCalc="1"' in generated_workbook

        reference_curr = reference.read("xl/worksheets/sheet1.xml").decode("utf-8")
        generated_curr = generated.read("xl/worksheets/sheet1.xml").decode("utf-8")
        assert _scrub_currency_cells(generated_curr) == _scrub_currency_cells(reference_curr)

        reference_kalk = reference.read("xl/worksheets/sheet2.xml").decode("utf-8")
        generated_kalk = generated.read("xl/worksheets/sheet2.xml").decode("utf-8")
        assert _scrub_calculator_cells(generated_kalk) == _scrub_calculator_cells(reference_kalk)


def _scrub_currency_cells(xml: str) -> str:
    xml = re.sub(r'<dimension\s+ref="[^"]+"\s*/>', '<dimension/>', xml, count=1)
    for row in range(2, 14):
        for column in ("B", "C"):
            xml = re.sub(
                rf'<c\b[^>]*?\br="{column}{row}"[^>]*?\s*(?:/>|>.*?</c>)',
                "",
                xml,
                flags=re.DOTALL,
            )
    xml = re.sub(r'<row\s+r="(?:8|9|10|11|12|13)"\s*>\s*</row>', "", xml)
    return xml


def _scrub_calculator_cells(xml: str) -> str:
    columns = [
        "B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P",
        "Q", "R", "S", "T", "U", "V", "W", "X", "Y", "Z", "AA", "AB", "AC", "AD",
        "AE", "AF", "AG", "AH", "AI", "AJ",
    ]
    refs = [f"{column}{row}" for row in range(DATA_START_ROW, DATA_END_ROW + 1) for column in columns]
    refs.extend([*TOTAL_FORMULAS, *PAYMENT_FORMULAS, "Z103", "Y104", "Z104"])
    for ref in refs:
        xml = re.sub(
            rf'<c\b[^>]*?\br="{ref}"[^>]*?\s*(?:/>|>.*?</c>)',
            "",
            xml,
            flags=re.DOTALL,
        )
    return xml
