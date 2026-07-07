"""Real Excel itinerary fixture bank for random product-output checks.

The source workbooks are intentionally kept as fixtures so Day Brain changes can
be tested against varied real supplier/calculator structures instead of only
small hand-written examples.
"""

from __future__ import annotations

import json
import random
import re
import sys
from dataclasses import dataclass
from datetime import date, datetime, time
from pathlib import Path
from typing import Any, Iterable, Iterator, Mapping, Sequence

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_FIXTURE_DIR = ROOT / "tests/fixtures/real_excel_inputs"
DEFAULT_MANIFEST = DEFAULT_FIXTURE_DIR / "manifest.json"
SKIP_SHEETS = {"curr", "overview", "oversikt", "ideas"}
HEADER_ALIASES = {
    "day": "day",
    "type": "type",
    "travel element": "travel_element",
    "from date": "from_date",
    "to date": "to_date",
    "from time": "from_time",
    "to time": "to_time",
}


@dataclass(frozen=True)
class WorkbookSpec:
    path: Path
    kind: str
    country_tags: tuple[str, ...]
    purpose_tags: tuple[str, ...]


@dataclass(frozen=True)
class ExcelFixtureCandidate:
    workbook_path: Path
    sheet_name: str
    kind: str
    country_tags: tuple[str, ...]
    purpose_tags: tuple[str, ...]
    row_count: int
    day_count: int
    raw_text: str

    @property
    def fixture_id(self) -> str:
        return f"{self.workbook_path.name}::{self.sheet_name}"

    @property
    def tags(self) -> tuple[str, ...]:
        sheet_tags = _tags_from_sheet_name(self.sheet_name)
        return tuple(dict.fromkeys((*self.country_tags, *self.purpose_tags, *sheet_tags)))

    def summary(self) -> dict[str, Any]:
        return {
            "fixture_id": self.fixture_id,
            "workbook": self.workbook_path.name,
            "sheet": self.sheet_name,
            "kind": self.kind,
            "row_count": self.row_count,
            "day_count": self.day_count,
            "tags": list(self.tags),
        }


def load_manifest(path: Path = DEFAULT_MANIFEST) -> tuple[WorkbookSpec, ...]:
    data = json.loads(path.read_text(encoding="utf-8"))
    specs: list[WorkbookSpec] = []
    for item in data.get("workbooks", []):
        workbook_path = (path.parent / str(item["file"])).resolve()
        specs.append(
            WorkbookSpec(
                path=workbook_path,
                kind=str(item.get("kind") or "unknown"),
                country_tags=tuple(str(tag) for tag in item.get("country_tags", ())),
                purpose_tags=tuple(str(tag) for tag in item.get("purpose_tags", ())),
            )
        )
    return tuple(specs)


def _clean_header(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").strip()).casefold()


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        if value.time() == time(0, 0):
            return value.strftime("%d.%m.%Y")
        return value.strftime("%d.%m.%Y %H:%M")
    if isinstance(value, date):
        return value.strftime("%d.%m.%Y")
    if isinstance(value, time):
        return value.strftime("%H:%M")
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return re.sub(r"\s+", " ", str(value)).strip()


def _find_header_map(sheet: Worksheet) -> tuple[int, Mapping[str, int]] | None:
    for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        values = [_clean_header(value) for value in row]
        header_map: dict[str, int] = {}
        for index, value in enumerate(values):
            canonical = HEADER_ALIASES.get(value)
            if canonical and canonical not in header_map:
                header_map[canonical] = index
        if {"day", "type", "travel_element"}.issubset(header_map):
            return row_index, header_map
        if row_index > 12:
            return None
    return None


def _row_has_itinerary_content(row_values: Sequence[str], header_map: Mapping[str, int]) -> bool:
    day = row_values[header_map["day"]] if len(row_values) > header_map["day"] else ""
    item_type = row_values[header_map["type"]] if len(row_values) > header_map["type"] else ""
    travel_element = row_values[header_map["travel_element"]] if len(row_values) > header_map["travel_element"] else ""
    return bool(day and item_type and travel_element)


def _tags_from_sheet_name(sheet_name: str) -> tuple[str, ...]:
    name = sheet_name.casefold()
    tags: list[str] = []
    if "sd" in name:
        tags.append("self_drive")
    if "gt" in name:
        tags.append("group_tour")
    if "winter" in name or "w" in name.split():
        tags.append("winter")
    if "summer" in name or "s" in name.split():
        tags.append("summer")
    if "reverse" in name:
        tags.append("reverse_route")
    return tuple(tags)


def extract_sheet_candidate(
    spec: WorkbookSpec,
    sheet: Worksheet,
    *,
    min_rows: int = 3,
    max_columns: int | None = None,
) -> ExcelFixtureCandidate | None:
    header = _find_header_map(sheet)
    if not header:
        return None
    header_row, header_map = header
    raw_lines: list[str] = []
    day_values: set[str] = set()
    itinerary_width = header_map["travel_element"] + 1
    width = min(max_columns or itinerary_width, sheet.max_column or itinerary_width)

    for row in sheet.iter_rows(min_row=header_row + 1, max_col=width, values_only=True):
        row_values = [_cell_text(value) for value in row]
        if not _row_has_itinerary_content(row_values, header_map):
            continue
        day_values.add(row_values[header_map["day"]])
        raw_lines.append("\t".join(row_values))

    if len(raw_lines) < min_rows:
        return None

    return ExcelFixtureCandidate(
        workbook_path=spec.path,
        sheet_name=sheet.title,
        kind=spec.kind,
        country_tags=spec.country_tags,
        purpose_tags=spec.purpose_tags,
        row_count=len(raw_lines),
        day_count=len(day_values),
        raw_text="\n".join(raw_lines),
    )


def iter_excel_fixture_candidates(
    manifest_path: Path = DEFAULT_MANIFEST,
    *,
    include_workbooks: Iterable[str] = (),
    min_rows: int = 3,
) -> Iterator[ExcelFixtureCandidate]:
    include_terms = tuple(term.casefold() for term in include_workbooks if term)
    for spec in load_manifest(manifest_path):
        if include_terms and not any(term in spec.path.name.casefold() for term in include_terms):
            continue
        workbook = load_workbook(spec.path, data_only=True, read_only=True)
        try:
            for sheet in workbook.worksheets:
                if sheet.title.casefold() in SKIP_SHEETS:
                    continue
                candidate = extract_sheet_candidate(spec, sheet, min_rows=min_rows)
                if candidate is not None:
                    yield candidate
        finally:
            workbook.close()


def build_candidate_index(manifest_path: Path = DEFAULT_MANIFEST) -> tuple[ExcelFixtureCandidate, ...]:
    return tuple(iter_excel_fixture_candidates(manifest_path))


def select_random_candidates(
    candidates: Sequence[ExcelFixtureCandidate],
    *,
    sample_size: int,
    seed: int,
    max_per_workbook: int = 2,
) -> tuple[ExcelFixtureCandidate, ...]:
    if sample_size <= 0:
        return ()
    rng = random.Random(seed)
    shuffled = list(candidates)
    rng.shuffle(shuffled)
    selected: list[ExcelFixtureCandidate] = []
    per_workbook: dict[str, int] = {}
    for candidate in shuffled:
        if len(selected) >= sample_size:
            break
        workbook_key = candidate.workbook_path.name
        if per_workbook.get(workbook_key, 0) >= max_per_workbook:
            continue
        selected.append(candidate)
        per_workbook[workbook_key] = per_workbook.get(workbook_key, 0) + 1
    if len(selected) < sample_size:
        for candidate in shuffled:
            if len(selected) >= sample_size:
                break
            if candidate not in selected:
                selected.append(candidate)
    return tuple(selected)


def write_candidate_raw_text(candidate: ExcelFixtureCandidate, output_dir: Path) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    safe_workbook = re.sub(r"[^A-Za-z0-9_.-]+", "_", candidate.workbook_path.stem)
    safe_sheet = re.sub(r"[^A-Za-z0-9_.-]+", "_", candidate.sheet_name).strip("_") or "sheet"
    output_path = output_dir / f"{safe_workbook}__{safe_sheet}.txt"
    output_path.write_text(candidate.raw_text + "\n", encoding="utf-8")
    return output_path


def build_index_summary(candidates: Sequence[ExcelFixtureCandidate]) -> dict[str, Any]:
    by_workbook: dict[str, int] = {}
    total_rows = 0
    for candidate in candidates:
        by_workbook[candidate.workbook_path.name] = by_workbook.get(candidate.workbook_path.name, 0) + 1
        total_rows += candidate.row_count
    return {
        "workbook_count": len(by_workbook),
        "candidate_count": len(candidates),
        "total_candidate_rows": total_rows,
        "candidates_by_workbook": dict(sorted(by_workbook.items())),
    }


if __name__ == "__main__":
    candidates = build_candidate_index()
    print(json.dumps(build_index_summary(candidates), ensure_ascii=False, indent=2))
