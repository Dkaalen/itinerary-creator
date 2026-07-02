"""Import Local Library seed rows from workbook files."""

from __future__ import annotations

from pathlib import Path
from typing import Iterable, Mapping
from zipfile import ZipFile
from xml.etree import ElementTree as ET

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from calculator.library_model import LOCAL_LIBRARY_HEADERS, LOCAL_LIBRARY_SHEET_NAME, LocalLibraryRow
from calculator.library_normalize import normalize_library_mapping, normalize_library_rows

_IGNORED_CHEAT_SHEETS = {"curr"}

_CLEAN_INPUTS_SHEET_NAME = "Clean Inputs"
_CLEAN_INPUT_BLOCKS = {
    "General Inputs": (1, 7),
    "Activities": (9, 15),
    "Group Tours": (20, 26),
}
_CLEAN_INPUT_HEADER_ROW = 2
_CLEAN_INPUT_DATA_START_ROW = 3
_XML_NS = {"x": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
_OFFICE_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"


def import_seed_workbook(path: str | Path) -> tuple[LocalLibraryRow, ...]:
    """Import rows from a Google-ready Local Library workbook or Cheat Sheet workbook."""

    workbook_path = Path(path)
    local_library_rows = import_local_library_workbook(workbook_path)
    if local_library_rows:
        return local_library_rows
    clean_input_rows = import_clean_inputs_workbook(workbook_path)
    if clean_input_rows:
        return clean_input_rows
    return import_cheat_sheet_workbook(workbook_path)


def import_local_library_workbook(
    path: str | Path,
    worksheet_name: str = LOCAL_LIBRARY_SHEET_NAME,
) -> tuple[LocalLibraryRow, ...]:
    """Import normalized rows from a Google-ready Local Library sheet."""

    raw_rows = _read_named_sheet_rows(Path(path), worksheet_name)
    if not raw_rows or not _looks_like_local_library(raw_rows[0].keys()):
        return ()
    return normalize_library_rows(raw_rows)


def import_clean_inputs_workbook(
    path: str | Path,
    *,
    blocks: Iterable[str] = ("General Inputs",),
) -> tuple[LocalLibraryRow, ...]:
    """Import rows from the Clean Inputs workbook's named source blocks."""

    workbook = load_workbook(Path(path), data_only=False)
    if _CLEAN_INPUTS_SHEET_NAME not in workbook.sheetnames:
        return ()
    sheet = workbook[_CLEAN_INPUTS_SHEET_NAME]
    selected_blocks = tuple(blocks)
    rows: list[LocalLibraryRow] = []
    for block_name in selected_blocks:
        if block_name not in _CLEAN_INPUT_BLOCKS:
            continue
        start_column, end_column = _CLEAN_INPUT_BLOCKS[block_name]
        for source_row, raw_row in _clean_input_block_rows(sheet, block_name, start_column, end_column):
            rows.append(normalize_library_mapping(raw_row | {"source_row": source_row}))
    return tuple(rows)


def import_cheat_sheet_workbook(
    path: str | Path,
    sheet_names: Iterable[str] | None = None,
) -> tuple[LocalLibraryRow, ...]:
    """Import calculator-like rows from Cheat Sheet source sheets."""

    workbook = load_workbook(Path(path), data_only=False)
    selected_names = tuple(sheet_names or workbook.sheetnames)
    rows: list[LocalLibraryRow] = []
    for sheet_name in selected_names:
        if sheet_name.lower() in _IGNORED_CHEAT_SHEETS or sheet_name not in workbook.sheetnames:
            continue
        raw_rows = _worksheet_rows(workbook[sheet_name])
        if not raw_rows:
            continue
        for raw_row in _cheat_sheet_rows(sheet_name, raw_rows):
            rows.append(normalize_library_mapping(raw_row))
    return tuple(rows)


def _read_named_sheet_rows(path: Path, worksheet_name: str) -> tuple[Mapping[str, object], ...]:
    rows = _read_openpyxl_named_sheet_rows(path, worksheet_name)
    if rows:
        return rows
    return _read_ooxml_named_sheet_rows(path, worksheet_name)


def _read_openpyxl_named_sheet_rows(path: Path, worksheet_name: str) -> tuple[Mapping[str, object], ...]:
    workbook = load_workbook(path, data_only=False, read_only=True)
    if worksheet_name not in workbook.sheetnames:
        return ()
    return _worksheet_rows(workbook[worksheet_name])


def _worksheet_rows(sheet: Worksheet) -> tuple[Mapping[str, object], ...]:
    values = tuple(sheet.iter_rows(values_only=True))
    return _rows_from_values(values)


def _rows_from_values(values: tuple[tuple[object, ...], ...]) -> tuple[Mapping[str, object], ...]:
    if not values:
        return ()
    header_index = _header_index(values)
    if header_index is None:
        return ()
    headers = tuple(str(value).strip() if value is not None else "" for value in values[header_index])
    rows = []
    for row in values[header_index + 1 :]:
        mapped = {header: value for header, value in zip(headers, row) if header}
        if any(value not in (None, "") for value in mapped.values()):
            rows.append(mapped)
    return tuple(rows)


def _header_index(values: tuple[tuple[object, ...], ...]) -> int | None:
    for index, row in enumerate(values[:10]):
        headers = {str(value).strip() for value in row if value is not None}
        if "Travel element" in headers or "library_id" in headers:
            return index
    return None


def _cheat_sheet_rows(sheet_name: str, raw_rows: tuple[Mapping[str, object], ...]) -> tuple[Mapping[str, object], ...]:
    rows = []
    for offset, raw_row in enumerate(raw_rows, start=1):
        travel_element = raw_row.get("Travel element")
        row_type = raw_row.get("Type")
        if not travel_element and not row_type:
            continue
        enriched = dict(raw_row)
        enriched.setdefault("schema_version", "local_library_v1")
        enriched.setdefault("source_workbook", "Cheat Sheet 2.0.xlsx")
        enriched.setdefault("source_sheet", sheet_name)
        enriched.setdefault("source_row", offset)
        enriched.setdefault("country", sheet_name)
        enriched.setdefault("category", row_type or sheet_name)
        enriched.setdefault("record_type", "line")
        enriched.setdefault("is_deleted", False)
        enriched.setdefault("is_fetchable", True)
        rows.append(enriched)
    return tuple(rows)


def _clean_input_block_rows(
    sheet: Worksheet,
    block_name: str,
    start_column: int,
    end_column: int,
) -> tuple[tuple[int, Mapping[str, object]], ...]:
    headers = [sheet.cell(_CLEAN_INPUT_HEADER_ROW, column).value for column in range(start_column, end_column + 1)]
    rows: list[tuple[int, Mapping[str, object]]] = []
    for row_number in range(_CLEAN_INPUT_DATA_START_ROW, sheet.max_row + 1):
        values = [sheet.cell(row_number, column).value for column in range(start_column, end_column + 1)]
        mapped = {str(header).strip(): value for header, value in zip(headers, values) if header}
        if not _clean_input_row_has_content(mapped):
            continue
        rows.append((row_number, _clean_input_mapping(block_name, mapped)))
    return tuple(rows)


def _clean_input_mapping(block_name: str, row: Mapping[str, object]) -> Mapping[str, object]:
    city = row.get("City / Area")
    details = row.get("Details")
    row_type = row.get("Type")
    price = row.get("Unit P") if row.get("Unit P") not in (None, "", 0) else row.get("Gross P")
    return {
        "schema_version": "local_library_v1",
        "source_workbook": "Clean Inputs.xlsx",
        "source_sheet": block_name,
        "country": _clean_input_country(block_name, city),
        "category": block_name,
        "record_type": "line",
        "is_deleted": False,
        "is_fetchable": True,
        "Type": row_type,
        "Supplier": "",
        "Travel element": details,
        "Comments": f"{block_name} · {city or ''}".strip(" ·"),
        "URL": row.get("URL"),
        "Gross P per unit": price,
        "Units": 1 if price not in (None, "", 0) else 0,
        "Supp Comm": row.get("Supp Margin"),
        "Supp curr": "NOK",
        "Sales curr": "NOK",
        "updated_by": "bundled_clean_inputs",
    }


def _clean_input_country(block_name: str, city: object) -> str:
    if block_name == "General Inputs":
        return "General Inputs"
    return str(city or block_name).strip()


def _clean_input_row_has_content(row: Mapping[str, object]) -> bool:
    return bool(str(row.get("Details") or "").strip() or str(row.get("Type") or "").strip())


def _looks_like_local_library(headers: Iterable[str]) -> bool:
    header_set = {str(header).strip() for header in headers}
    required = {"schema_version", "library_id", "is_deleted", "is_fetchable", "Travel element"}
    return required.issubset(header_set) and len(header_set.intersection(LOCAL_LIBRARY_HEADERS)) >= 20


def _read_ooxml_named_sheet_rows(path: Path, worksheet_name: str) -> tuple[Mapping[str, object], ...]:
    with ZipFile(path) as archive:
        sheet_path = _worksheet_xml_path(archive, worksheet_name)
        if not sheet_path:
            return ()
        root = ET.fromstring(archive.read(sheet_path))
        values = tuple(_xml_row_values(row) for row in root.findall(".//x:sheetData/x:row", _XML_NS))
    return _rows_from_values(values)


def _worksheet_xml_path(archive: ZipFile, worksheet_name: str) -> str:
    workbook = ET.fromstring(archive.read("xl/workbook.xml"))
    rels = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
    rel_by_id = {rel.attrib["Id"]: rel.attrib["Target"] for rel in rels}
    for sheet in workbook.findall(".//x:sheet", _XML_NS):
        if sheet.attrib.get("name") != worksheet_name:
            continue
        rel_id = sheet.attrib.get(f"{{{_OFFICE_REL_NS}}}id", "")
        target = rel_by_id.get(rel_id, "")
        return target.lstrip("/") if target.startswith("/xl/") else f"xl/{target.lstrip('/')}"
    return ""


def _xml_row_values(row: ET.Element) -> tuple[object, ...]:
    values = []
    expected_column = 1
    for cell in row.findall("x:c", _XML_NS):
        column = _column_number(cell.attrib.get("r", ""))
        while expected_column < column:
            values.append(None)
            expected_column += 1
        values.append(_xml_cell_value(cell))
        expected_column += 1
    return tuple(values)


def _xml_cell_value(cell: ET.Element) -> object:
    value = cell.find("x:v", _XML_NS)
    if value is not None:
        return value.text
    inline_string = cell.find("x:is", _XML_NS)
    if inline_string is None:
        return None
    return "".join(text.text or "" for text in inline_string.iter(f"{{{_XML_NS['x']}}}t"))


def _column_number(reference: str) -> int:
    letters = "".join(char for char in reference if char.isalpha())
    number = 0
    for char in letters:
        number = number * 26 + ord(char.upper()) - 64
    return number or 1
