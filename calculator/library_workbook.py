"""Validated repository-local Excel Local Library loader."""
from __future__ import annotations

from functools import lru_cache
from pathlib import Path

from openpyxl import load_workbook

from calculator.library_normalize import LocalLibraryNumericValueError, normalize_library_mapping
from calculator.library_workbook_diagnostics import diagnostic as _diagnostic
from calculator.library_workbook_diagnostics import display_value as _display_value
from calculator.library_workbook_formulas import (
    canonical_formula_syntax_is_plausible as _canonical_formula_syntax_is_plausible,
    formula_and_cached_value_issues as _formula_and_cached_value_issues,
    formula_syntax_is_plausible as _formula_syntax_is_plausible,
    read_formula_cells as _read_formula_cells,
    read_sheet_formula_cells as _read_sheet_formula_cells,
)
from calculator.library_workbook_models import (
    FormulaCell as _FormulaCell,
    LocalLibraryDiagnostic,
    LocalLibraryWorkbook,
    LocalLibraryWorkbookError,
)
from calculator.library_workbook_rows import (
    has_product_content as _has_product_content,
    row_validation_issue as _row_validation_issue,
)
from calculator.library_workbook_schema import (
    DATA_SHEETS,
    REQUIRED_HEADERS as _REQUIRED_HEADERS,
    REQUIRED_SHEETS,
    currency_rates as _currency_rates,
    headers as _headers,
    validate_required_sheets as _validate_required_sheets,
)

WORKBOOK_PATH = (
    Path(__file__).resolve().parents[1]
    / "data"
    / "Calculation-template-Inputs-fixed-outline-restored.xlsx"
)


@lru_cache(maxsize=4)
def _load_cached(path_text: str, modified_ns: int, size: int) -> LocalLibraryWorkbook:
    path = Path(path_text)
    workbook = None
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
        _validate_required_sheets(workbook.sheetnames)
        formula_cells = _read_formula_cells(path, workbook)

        rates, currency_diagnostics = _currency_rates(workbook["Curr"])
        diagnostics: list[LocalLibraryDiagnostic] = list(currency_diagnostics)
        rows = []

        for sheet_name in DATA_SHEETS:
            sheet = workbook[sheet_name]
            headers, header_row = _headers(sheet)
            missing_headers = sorted(_REQUIRED_HEADERS - set(headers))
            if missing_headers:
                raise LocalLibraryWorkbookError(
                    f"{sheet_name} is missing required header(s): {', '.join(missing_headers)}",
                    category="schema",
                    code="missing_headers",
                )

            for source_row, value_cells in enumerate(
                sheet.iter_rows(min_row=header_row + 1),
                start=header_row + 1,
            ):
                values = tuple(cell.value for cell in value_cells)
                mapping = {header: value for header, value in zip(headers, values) if header}
                if not _has_product_content(mapping):
                    continue

                row_issues = _formula_and_cached_value_issues(
                    headers,
                    formula_cells.get(sheet_name, {}).get(source_row, ()),
                    sheet_name,
                    source_row,
                )
                diagnostics.extend(row_issues)
                if any(issue.category == "invalid_record" for issue in row_issues):
                    continue

                enriched = dict(mapping)
                enriched.update(
                    {
                        "schema_version": "local_library_v1",
                        "source_workbook": path.name,
                        "source_sheet": sheet_name,
                        "source_row": source_row,
                        "country": str(mapping.get("ID") or "").strip(),
                        "category": sheet_name,
                        "record_type": "line",
                        "is_deleted": False,
                        "is_fetchable": True,
                    }
                )
                try:
                    row = normalize_library_mapping(enriched, strict_numeric=True)
                except LocalLibraryNumericValueError as error:
                    diagnostics.append(
                        _diagnostic(
                            category="invalid_record",
                            code="invalid_numeric_value",
                            worksheet=sheet_name,
                            excel_row=source_row,
                            field=error.source_field or error.field_name,
                            value=error.value,
                            reason=error.reason,
                        )
                    )
                    continue

                validation_issue = _row_validation_issue(row, sheet_name, source_row, rates)
                if validation_issue is not None:
                    diagnostics.append(validation_issue)
                    continue
                rows.append(row)

        if not rows:
            raise LocalLibraryWorkbookError(
                "Local Library workbook contains no fetchable rows.",
                category="fatal_workbook",
                code="no_fetchable_rows",
                diagnostics=tuple(diagnostics),
            )
        fingerprint = f"{modified_ns:x}-{size:x}"
        return LocalLibraryWorkbook(tuple(rows), rates, path, fingerprint, tuple(diagnostics))
    except LocalLibraryWorkbookError:
        raise
    except Exception as exc:
        raise LocalLibraryWorkbookError(
            f"Local Library workbook is corrupt or unreadable: {path.name}",
            category="fatal_workbook",
            code="unreadable_workbook",
        ) from exc
    finally:
        if workbook is not None:
            workbook.close()


def load_local_library_workbook(path: str | Path = WORKBOOK_PATH) -> LocalLibraryWorkbook:
    workbook_path = Path(path)
    if not workbook_path.is_file():
        raise LocalLibraryWorkbookError(
            f"Local Library workbook is missing: {workbook_path}",
            category="fatal_workbook",
            code="missing_workbook",
        )
    stat = workbook_path.stat()
    return _load_cached(str(workbook_path.resolve()), stat.st_mtime_ns, stat.st_size)


def clear_local_library_workbook_cache() -> None:
    _load_cached.cache_clear()


__all__ = [
    "DATA_SHEETS",
    "REQUIRED_SHEETS",
    "WORKBOOK_PATH",
    "LocalLibraryDiagnostic",
    "LocalLibraryWorkbook",
    "LocalLibraryWorkbookError",
    "clear_local_library_workbook_cache",
    "load_local_library_workbook",
]
