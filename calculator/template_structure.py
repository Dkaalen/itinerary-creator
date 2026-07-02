"""Read and validate the calculation workbook structure."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter

from calculator.columns import (
    ADVANCED_COLUMN_RANGES,
    AUTO_FILTER_REF,
    COLUMN_SPECS,
    CURRENCY_SHEET_NAME,
    DATA_END_ROW,
    DATA_START_ROW,
    HEADER_BY_LETTER,
    HEADER_ROW,
    KALK_SHEET_NAME,
    PAYMENT_END_ROW,
    PAYMENT_START_ROW,
    TOTALS_ROW,
    WORKBOOK_END_COLUMN,
    WORKBOOK_START_COLUMN,
)

_PACKAGE_DIR = Path(__file__).resolve().parent
DEFAULT_TEMPLATE_PATH = _PACKAGE_DIR / "templates" / "Calculation-template-Mal.xlsx"


@dataclass(frozen=True)
class TemplateStructure:
    """Locked workbook layout needed by the calculator exporter."""

    sheet_names: tuple[str, ...]
    active_sheet: str
    kalk_max_row: int
    kalk_max_column: int
    header_row: int
    data_start_row: int
    data_end_row: int
    totals_row: int
    payment_start_row: int
    payment_end_row: int
    workbook_start_column: str
    workbook_end_column: str
    auto_filter_ref: str | None
    headers_by_letter: dict[str, str]
    hidden_column_ranges: tuple[str, ...]
    grouped_column_ranges: tuple[str, ...]
    collapsed_column_markers: tuple[str, ...]
    currencies: dict[str, float]


def default_template_path() -> Path:
    """Return the bundled calculation template path."""

    return DEFAULT_TEMPLATE_PATH


def inspect_template_structure(path: str | Path | None = None) -> TemplateStructure:
    """Inspect the template workbook without modifying it."""

    template_path = Path(path) if path is not None else default_template_path()
    workbook = load_workbook(template_path, data_only=False)
    kalk_sheet = workbook[KALK_SHEET_NAME]
    currency_sheet = workbook[CURRENCY_SHEET_NAME]

    return TemplateStructure(
        sheet_names=tuple(workbook.sheetnames),
        active_sheet=workbook.active.title,
        kalk_max_row=kalk_sheet.max_row,
        kalk_max_column=kalk_sheet.max_column,
        header_row=HEADER_ROW,
        data_start_row=DATA_START_ROW,
        data_end_row=DATA_END_ROW,
        totals_row=TOTALS_ROW,
        payment_start_row=PAYMENT_START_ROW,
        payment_end_row=PAYMENT_END_ROW,
        workbook_start_column=WORKBOOK_START_COLUMN,
        workbook_end_column=WORKBOOK_END_COLUMN,
        auto_filter_ref=kalk_sheet.auto_filter.ref,
        headers_by_letter=_read_headers(kalk_sheet),
        hidden_column_ranges=_column_ranges(kalk_sheet, "hidden"),
        grouped_column_ranges=_column_ranges(kalk_sheet, "outlineLevel"),
        collapsed_column_markers=_collapsed_markers(kalk_sheet),
        currencies=_read_currencies(currency_sheet),
    )


def validate_template_structure(path: str | Path | None = None) -> tuple[str, ...]:
    """Return human-readable structure issues for the template."""

    structure = inspect_template_structure(path)
    issues: list[str] = []
    if structure.sheet_names != (CURRENCY_SHEET_NAME, KALK_SHEET_NAME):
        issues.append(f"Expected sheets Curr/Kalk, got {structure.sheet_names!r}.")
    if structure.active_sheet != KALK_SHEET_NAME:
        issues.append(f"Expected active sheet Kalk, got {structure.active_sheet!r}.")
    if structure.auto_filter_ref != AUTO_FILTER_REF:
        issues.append(f"Expected auto-filter {AUTO_FILTER_REF}, got {structure.auto_filter_ref!r}.")
    if structure.headers_by_letter != HEADER_BY_LETTER:
        issues.append("Kalk header row does not match calculator columns.")
    if structure.hidden_column_ranges != ADVANCED_COLUMN_RANGES:
        issues.append(f"Expected hidden ranges {ADVANCED_COLUMN_RANGES}, got {structure.hidden_column_ranges!r}.")
    if structure.grouped_column_ranges != ADVANCED_COLUMN_RANGES:
        issues.append(f"Expected grouped ranges {ADVANCED_COLUMN_RANGES}, got {structure.grouped_column_ranges!r}.")
    if not structure.currencies:
        issues.append("Curr sheet has no currency lookup values.")
    return tuple(issues)


def _read_headers(kalk_sheet: Any) -> dict[str, str]:
    return {
        column.letter: kalk_sheet[f"{column.letter}{HEADER_ROW}"].value
        for column in COLUMN_SPECS
    }


def _read_currencies(currency_sheet: Any) -> dict[str, float]:
    currencies: dict[str, float] = {}
    for row in range(2, 14):
        code = currency_sheet[f"B{row}"].value
        rate = currency_sheet[f"C{row}"].value
        if code is None or rate is None:
            continue
        currencies[str(code)] = float(rate)
    return currencies


def _column_ranges(sheet: Any, attribute: str) -> tuple[str, ...]:
    ranges: list[str] = []
    for letter in sorted(sheet.column_dimensions, key=_column_sort_key):
        dimension = sheet.column_dimensions[letter]
        value = getattr(dimension, attribute)
        is_active = bool(value) if attribute == "hidden" else int(value or 0) > 0
        if not is_active:
            continue
        start = int(dimension.min or _column_sort_key(letter))
        end = int(dimension.max or start)
        ranges.append(_format_column_range(start, end))
    return tuple(ranges)


def _collapsed_markers(sheet: Any) -> tuple[str, ...]:
    markers: list[str] = []
    for letter in sorted(sheet.column_dimensions, key=_column_sort_key):
        if not sheet.column_dimensions[letter].collapsed:
            continue
        if _is_inside_workbook_columns(letter):
            markers.append(letter)
    return tuple(markers)


def _is_inside_workbook_columns(letter: str) -> bool:
    column_index = column_index_from_string(letter)
    start_index = column_index_from_string(WORKBOOK_START_COLUMN)
    end_index = column_index_from_string(WORKBOOK_END_COLUMN)
    return start_index <= column_index <= end_index


def _format_column_range(start: int, end: int) -> str:
    first = get_column_letter(start)
    last = get_column_letter(end)
    return f"{first}:{last}"


def _column_sort_key(letter: str) -> int:
    number = 0
    for char in letter:
        number = number * 26 + ord(char.upper()) - ord("A") + 1
    return number
